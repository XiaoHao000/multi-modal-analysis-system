"""
Excel 表格处理器: .xlsx / .xls 结构化数据提取。

流程:
  .xlsx → openpyxl 逐 sheet 读取 → list[dict]
  .xls  → xlrd 逐 sheet 读取 → list[dict]

每个 sheet 返回为:
  {"sheet_name": "Sheet1", "columns": [...], "rows": [[...], ...], "row_count": N}
"""

import base64
import io
from typing import List

from utils.logger import logger


def _decode_to_bytes(data_b64: str) -> bytes:
    return base64.b64decode(data_b64)


def _extract_xlsx(file_bytes: bytes) -> List[dict]:
    """从 .xlsx 文件中提取所有 sheet 的结构化数据。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        columns = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
        data_rows = []
        for row in rows[1:]:
            data_rows.append([_clean_cell(v) for v in row])
        sheets.append({
            "sheet_name": sheet_name,
            "columns": columns,
            "rows": data_rows[:500],  # 每 sheet 最多 500 行
            "row_count": len(data_rows),
        })
    wb.close()
    return sheets


def _extract_xls(file_bytes: bytes) -> List[dict]:
    """从 .xls 文件中提取所有 sheet 的结构化数据。"""
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sheets = []
    for sheet in wb.sheets():
        rows = []
        for r in range(min(sheet.nrows, 501)):
            rows.append([_clean_cell(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
        if not rows:
            continue
        columns = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
        data_rows = rows[1:]
        sheets.append({
            "sheet_name": sheet.name,
            "columns": columns,
            "rows": data_rows,
            "row_count": sheet.nrows - 1,
        })
    return sheets


def _clean_cell(value):
    """清洗单元格值：NaN → None，数字保留合理精度。"""
    import math
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return round(value, 4)
    if isinstance(value, str):
        return value.strip()
    return value


async def process_excel_files(files: List[dict], user_query: str = "") -> List[dict]:
    """处理 Excel 文件：提取结构化表格数据。

    Args:
        files: Excel 文件列表
        user_query: 用户原始问题（保留接口一致，当前未使用）

    Returns:
        excel_data: [{"sheet_name": str, "columns": [...], "rows": [[...], ...], "row_count": N}, ...]
    """
    if not files:
        return []

    all_sheets = []

    for entry in files:
        data = entry.get("data", "")
        filename = entry.get("filename", "unknown")
        mime = entry.get("mime_type", "")
        if not data:
            continue

        try:
            file_bytes = _decode_to_bytes(data)
        except Exception as e:
            logger.warning(f"Excel base64 解码失败 ({filename}): {e}")
            continue

        try:
            if "openxmlformats" in mime or filename.lower().endswith(".xlsx"):
                sheets = _extract_xlsx(file_bytes)
            else:
                sheets = _extract_xls(file_bytes)
        except Exception as e:
            logger.warning(f"Excel 解析失败 ({filename}): {e}")
            sheets = []

        all_sheets.extend(sheets)
        total_rows = sum(s["row_count"] for s in sheets)
        logger.info(
            f"Excel 处理完成: {filename} → {len(sheets)} sheet, {total_rows} 行"
        )

    return all_sheets
